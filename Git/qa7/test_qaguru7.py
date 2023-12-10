from pathlib import Path
from zipfile import ZipFile
from openpyxl import load_workbook
import pytest
from PyPDF2 import PdfReader
import csv
import os.path

current_dir = Path(__file__).resolve().parent
resources_dir = f"{current_dir}/resources"
zip_file_path = os.path.join(resources_dir, 'archive.zip')
pdf_file_name = 'sample1.pdf'
xlsx_file_name = 'sample2.xlsx'
csv_file_name = 'sample3.csv'


@pytest.fixture
def archive_file():
    """Check if the archive exists"""
    if not os.path.exists(zip_file_path):

        files_to_zip = [
            os.path.join(resources_dir, 'sample1.pdf'),
            os.path.join(resources_dir, 'sample2.xlsx'),
            os.path.join(resources_dir, 'sample3.csv'),
        ]

        with ZipFile(zip_file_path, 'w') as zipf:
            # Add each file to the archive
            for file_path in files_to_zip:
                sample = os.path.basename(file_path)
                zipf.write(file_path, sample)

        print(f"Архив создан по пути: {zip_file_path}")

    with ZipFile(zip_file_path, 'r') as archive:
        yield archive


def test_archive_file_exists(archive_file):
    """Check that files exist in the archive"""
    file_list = archive_file.namelist()
    assert file_list is not None


def test_check_pdf_content(archive_file):
    """Check PDF content"""
    with archive_file.open(pdf_file_name) as pdf_data:
        reader = PdfReader(pdf_data)
        assert "Как это сделать вы узнаете в следующем видео." in reader.pages[1].extract_text()


def test_check_xlsx_content(archive_file):
    """Check XLSX content"""
    with archive_file.open(xlsx_file_name) as xlsx_data:
        workbook = load_workbook(xlsx_data)
        sheet = workbook.active
        assert sheet.cell(row=2, column=2).value is not None


def test_check_csv_content(archive_file):
    """Check CSV content"""
    with archive_file.open(csv_file_name) as csvfile:
        csv_reader = csv.reader(csvfile, delimiter=',')
        assert csv_reader is not None
