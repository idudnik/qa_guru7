from openpyxl import load_workbook
import pytest
from zipfile import ZipFile
from PyPDF2 import PdfReader
import csv

zip_file_path = '/Users/stanislavdudnik/Desktop/Git/qa_guru7/resources/archive.zip'
pdf_file_name = 'sample1.pdf'
xlsx_file_name = 'sample2.xlsx'
csv_file_name = 'sample3.csv'


@pytest.fixture
def archive_file():
    with ZipFile(zip_file_path, 'r') as archive:
        yield archive  # Yield the ZipFile object


def test_archive_file_exists(archive_file):
    """Проверяем, что в архиве есть файлы"""
    file_list = archive_file.namelist()
    assert file_list is not None


def test_check_pdf_content(archive_file):
    """Проверяем, пдф"""
    with archive_file.open(pdf_file_name) as pdf_data:
        reader = PdfReader(pdf_data)

        assert "Как это сделать вы узнаете в следующем видео." in reader.pages[1].extract_text()


def test_check_xlsx_content(archive_file):
    """Проверяем XLSX"""
    with archive_file.open(xlsx_file_name) as xlsx_data:
        workbook = load_workbook(xlsx_data)
        sheet = workbook.active

        assert sheet.cell(row=2, column=2).value != None


def test_check_csv_content(archive_file):
    """Проверяем CSV"""
    with archive_file.open(csv_file_name) as csvfile:
        csv_reader = csv.reader(csvfile, delimiter=',')

        assert csv_reader is not None
